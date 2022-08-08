import os
import re
import sys
from PyQt5 import uic
from PyQt5 import sip
from PyQt5 import QtCore, QtWidgets, QtGui
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QIcon, QPixmap
from datetime import datetime, date
from openpyxl import load_workbook
from config import db
from images import images_store
from work_with_excel import ExcelWork as excel
from work_with_excel import ExcelMontage as monatge_excel


class ChooseTypeOfMenu(QMainWindow):
    # Выбор администрирования
    def __init__(self, parent=None, flag=Qt.Window):
        super().__init__(parent, flag)  # Call the inherited classes __init__ method
        uic.loadUi('ui/options/main_menu_choose.ui', self)
        self.add_data_to_box()
        self.closeButton.clicked.connect(self.exit_btn)
        self.saveButton.clicked.connect(self.save_choose_menu)
        self.option_menu = int(re.search('\d+', str(db.get_menu_main())).group(0))

    def add_data_to_box(self):
        # Подгрузка комбо-боксов в меню в зависимости от выбора в настройках (который хранится в базе даных)
        self.type_menuBox.clear()
        self.option_menu = int(re.search('\d+', str(db.get_menu_main())).group(0))
        if self.option_menu == 1:
            self.type_menuBox.addItem("Настройка")
            self.type_menuBox.addItem("Производственное подразделение")
            self.type_menuBox.addItem("Монтажное подразделение")

        elif self.option_menu == 2:
            self.type_menuBox.addItem("Производственное подразделение")
            self.type_menuBox.addItem("Монтажное подразделение")
            self.type_menuBox.addItem("Настройка")
        elif self.option_menu == 3:
            self.type_menuBox.addItem("Монтажное подразделение")
            self.type_menuBox.addItem("Производственное подразделение")
            self.type_menuBox.addItem("Настройка")

    def exit_btn(self):
        self.close()

    def save_choose_menu(self):
        # Сохранение меню в базе данных
        choose_menu = self.type_menuBox.currentText()
        if choose_menu == 'Общий':
            db.change_type_of_menu_common()
            self.close()

        elif choose_menu == "Производственное подразделение":
            db.change_type_of_menu_produce()
            self.close()
            self.MainWindow = Main_menu()
            self.MainWindow.show()

        elif choose_menu == "Монтажное подразделение":
            db.change_type_of_menu_montage()
            self.close()
            self.MainWindow = Main_menu()
            self.MainWindow.show()


class Main_menu(QMainWindow):
    # Главное меню
    def __init__(self, parent=None, flag=Qt.Window):
        super().__init__(parent, flag)  # Call the inherited classes __init__ method optionsButton
        uic.loadUi('ui/main_manu_report.ui', self)
        # Загрузка меню в зависиомсти от выбора настроек
        self.option_menu = int(re.search('\d+', str(db.get_menu_main())).group(0))
        if self.option_menu == 2:
            # Подключение эвентов к кнопкам в зависимости от выброного меню
            self.access_to_bdButton.clicked.connect(self.access_butn_push)
            self.add_dataButton.clicked.connect(self.add_data_butn_push)
            self.reportButton.clicked.connect(self.report_butn_push)

        elif self.option_menu == 3:
            # Подключение эвентов к кнопкам в зависимости от выброного меню
            self.access_to_bdButton.clicked.connect(self.access_butn_push)
            self.add_dataButton.clicked.connect(self.montage_butn_push)
            self.reportButton.clicked.connect(self.montage_report_butn_push)

        self.optionsButton.clicked.connect(self.options_btn_press)
        self.MainWindow = ""

    def access_butn_push(self):
        # Доступ к меню выбра таблиц
        self.close()
        self.MainWindow = ChooseTable()
        self.MainWindow.show()

    def add_data_butn_push(self):
        # Добавление новых данных
        self.close()
        self.MainWindow = AddData()
        self.MainWindow.show()

    def report_butn_push(self):
        # Кнопка вывода отчета
        self.close()
        self.MainWindow = PrintReport()
        self.MainWindow.show()

    def montage_butn_push(self):
        # Кнопка добавление данных по монтажу
        self.close()
        self.MainWindow = AddMontageData()
        self.MainWindow.show()

    def montage_report_butn_push(self):
        # Кнопка вывода отчета по монтажу
        self.close()
        self.MainWindow = PrintReport()
        self.MainWindow.show()

    def options_btn_press(self):
        # Кнопка настроек меню
        self.close()
        self.MainWindow = ChooseTypeOfMenu()
        self.MainWindow.show()


class ChooseTable(QMainWindow):
    # Меню выбора таблиц
    def __init__(self, parent=None, flag=Qt.Window):
        super().__init__(parent, flag)
        uic.loadUi('ui/table_choose.ui', self)
        # Подгрузка меню в зависиомсти от меню выбранного
        self.option_menu = int(re.search('\d+', str(db.get_menu_main())).group(0))
        if self.option_menu == 2:
            self.table_1Button.clicked.connect(self.progect_table_btn)
            self.table_planButton.clicked.connect(self.table_data_plan_for_mounth)
            self.table_2Button.clicked.connect(self.table_data_ab_progect_btn)
            self.backButton.clicked.connect(self.back_btn_push)

        elif self.option_menu == 3:
            self.table_1Button.clicked.connect(self.montage_progect_table_btn)
            self.table_planButton.clicked.connect(self.table_data_plan_for_mounth)
            self.table_2Button.clicked.connect(self.table_data_ab_progect_btn)
            self.backButton.clicked.connect(self.back_btn_push)
        self.MainWindow = ""

    def progect_table_btn(self):
        # Таблица проектов
        self.close()
        self.MainWindow = ProgTable()
        self.MainWindow.show()

    def table_data_plan_for_mounth(self):
        # Таблица месячного плана
        self.close()
        self.MainWindow = PlanForMounth()
        self.MainWindow.show()

    def table_data_ab_progect_btn(self):
        # Таблица месячного плана
        self.close()
        self.MainWindow = ChooseProgect()
        self.MainWindow.show()

    def montage_progect_table_btn(self):
        # Таблица проектов для монтажников
        self.close()
        self.MainWindow = MontageProgTable()
        self.MainWindow.show()

    def back_btn_push(self):
        # Кнопка вернутся в главное меню
        self.close()
        self.MainWindow = Main_menu()
        self.MainWindow.show()


class ProgTable(QMainWindow):
    # Табица проектов
    def __init__(self, parent=None, flag=Qt.Window):
        super().__init__(parent, flag)
        uic.loadUi('ui/progect_name_table.ui', self)
        # Установим размеры колонок
        self.maintable.setColumnWidth(0, 250)
        for c in range(1, 4):
            self.maintable.setColumnWidth(c, 50)
        self.load_data_for_table()  # Загрузка данных в таблицу
        self.maintable.setWordWrap(True)  # Установим перенос слов в таблице
        self.add_str.triggered.connect(self.add_str_btn)  # Кнопка добавить позицию в таблицу
        self.edit_str.triggered.connect(self.edit_str_btn)  # Кнопка отредактироваь позицию в таблице
        self.refresh_1.triggered.connect(self.refresh_btn)  # Кнопка обновить данные
        self.delite_str.triggered.connect(self.delite_str_btn)  # Кнопка удалить позицию из таблицу
        self.exit_from_table.triggered.connect(self.exit_btn)  # Кнопка выхода
        self.MainWindow = ""

    def load_data_for_table(self):
        # Подгрузка данных из дб
        def set_row_in_table(row_from_db):
            # Метод для выравнивания значений в таблице по горизонтали и вертикали
            item = QtWidgets.QTableWidgetItem(row_from_db)
            item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
            return item

        data_prod = db.get_all_info_progect()
        self.maintable.setRowCount(len(data_prod))  # Устанавливаем количество колонок в таблице
        tablerow = 0
        # Назначаем информацию из базы данных в строки таблцы
        for row in db.get_all_info_progect():
            # Название проекта в 1 столбец устанавливаем
            self.maintable.setItem(tablerow, 0, QtWidgets.QTableWidgetItem(row[1]))
            col = 1
            # Показываем какой из проектов активен, где V - активен, Х - нет
            for index in range(2, 5):
                if row[index] is True or row[index] == 1:
                    self.maintable.setItem(tablerow, col, set_row_in_table('V'))
                else:
                    self.maintable.setItem(tablerow, col, set_row_in_table('X'))
                col += 1
            tablerow += 1

    def add_str_btn(self):
        # Кнопка добавления позиции в таблицу
        self.MainWindow = DialogAdd()
        self.MainWindow.show()

    def edit_str_btn(self):
        # Кнопка редактирования проекта
        if self.maintable.selectedIndexes():
            row = self.maintable.currentIndex().row()
            for i in self.maintable.selectedIndexes():
                # получаем значение строки из 1 стобца(наименование продукции)
                prog = self.maintable.model().index(row, 0).data()  # get cell at row, col
                data = [prog]
                self.MainWindow = DialogEdit(data)
                self.MainWindow.show()
        else:
            row = self.maintable.rowCount() - 1

    def refresh_btn(self):
        # Обновляем данные в таблице
        self.load_data_for_table()

    def delite_str_btn(self):
        # Удаление проекта
        progect = ''
        if self.maintable.selectedIndexes():
            row = self.maintable.currentIndex().row()
            progect = self.maintable.model().index(row, 0).data()  # get cell at row, col
        else:
            row = self.maintable.rowCount() - 1

        if row >= 0:
            title = "Подтверждение удаления позиции"
            text = f"Вы уверены, что хотите удалить данную позицию?\n"
            msg = MessageDialogWindow(title, text)  # Вызываем текстовое сообщение с подтверждением удаления
            if msg.conirm_message() == 1:
                self.maintable.removeRow(row)
                db.delite_prod(str(progect))
            else:
                self.update()

    def exit_btn(self):
        self.close()
        self.MainWindow = Main_menu()
        self.MainWindow.show()


class DialogAdd(QDialog):
    # Добавление проекта в базу данных
    def __init__(self, parent=None, flag=Qt.Dialog):
        super().__init__(parent, flag)
        uic.loadUi('ui/db actions/add_dialog.ui', self)
        self.add_btn_2.clicked.connect(self.add_prod_btn)

    def add_prod_btn(self):
        # Кнопка добавления позиций в таблицу
        prog = self.textEdit_prog.toPlainText()  # Берем текст из текстового виджета
        db.add_progect_name(prog)  # Добавляем в базу данных
        # Обновляем столбцы в дб активных проектов согласно выбора чекпоинтов
        if self.kmd_checkBox.isChecked():
            db.update_active_kmd(prog)
        else:
            db.update_disactive_kmd(prog)
        if self.spu_checkBox.isChecked():
            db.update_active_spu(prog)
        else:
            db.update_disactive_spu(prog)
        if self.pvh_checkBox.isChecked():
            db.update_active_pvh(prog)
        else:
            db.update_disactive_pvh(prog)
        if self.spu_checkBox.isChecked() and self.pvh_checkBox.isChecked():
            db.update_active_tent(prog)
        # Выводим сообщение об успехе
        message_title = "Успешное добавление"
        message_text = f"Проект успешно добавлен в базу данных."
        msg = MessageDialogWindow(message_title, message_text)
        if msg.success_msg() == 1:
            self.close()
            self.close()


class DialogEdit(QDialog):
    # Редактирование позиции проекта
    def __init__(self, data, parent=None, flag=Qt.Dialog):
        super().__init__(parent, flag)
        uic.loadUi('ui/db actions/edit_dialog.ui', self)
        self.change_btn.clicked.connect(self.change_prog_btn)  # Кнопка изменить
        self.data = data  # Полученная информация из строки в таблице
        progect = ''.join(self.data[0])  # Устанавливаем название проекта
        self.textEdit_prog.setText(f"{progect}")  # Устанавливаем название проекта
        self.id = str(int(re.search('\d+', str(db.get_id_progect(progect))).group(0)))  # Получаем id проекта из БД
        # Устанавливаем флаги в чекбоксах согласно данных из БД
        for p in db.get_info_about_this_progect(progect):
            if p[2] is True or p[2] == 1:
                self.kmd_checkBox.setChecked(True)
            if p[3] is True or p[3] == 1:
                self.spu_checkBox.setChecked(True)
            if p[4] is True or p[4] == 1:
                self.pvh_checkBox.setChecked(True)

    def change_prog_btn(self):
        # Кнопка Редактирования позиции проекта
        try:
            # Обновляем данные в БД
            progect = self.textEdit_prog.toPlainText()
            db.update_progect_name(progect, int(self.id))
            if self.kmd_checkBox.isChecked():
                db.update_active_kmd(progect)
            else:
                db.update_disactive_kmd(progect)
            if self.spu_checkBox.isChecked():
                db.update_active_spu(progect)
            else:
                db.update_disactive_spu(progect)
            if self.pvh_checkBox.isChecked():
                db.update_active_pvh(progect)
            else:
                db.update_disactive_pvh(progect)
            if self.spu_checkBox.isChecked() and self.pvh_checkBox.isChecked():
                db.update_active_tent(progect)
            # Выводим сообщение об успехе
            message_title = "Успешное изменение"
            message_text = f"Данные проекта успешно изменены."
            msg = MessageDialogWindow(message_title, message_text)
            if msg.success_msg() == 1:
                self.close()
        except:
            self.MainWindow = ErrorAddWin()
            self.MainWindow.show()


class PlanForMounth(QMainWindow):
    # Таблица с плановыми значениями
    def __init__(self, parent=None, flag=Qt.Window):
        super().__init__(parent, flag)
        uic.loadUi('ui/plan_table.ui', self)
        self.load_data_for_table()  # Загружаем данные в таблицу
        self.maintable.setWordWrap(True)  # Устанавливаем доступный перенос слов согласно содержания
        # Устанавливаем размер столбцов в таблице
        self.maintable.setColumnWidth(0, 55)
        for col in range(1, 4):
            self.maintable.setColumnWidth(col, 130)
        # Прописываем назначение кнопок
        self.add_str.triggered.connect(self.add_str_btn)  # Кнопка добавления данных в таблицу
        self.edit_str.triggered.connect(self.edit_str_btn)  # Кнопка редактирования данных таблицы
        self.refresh_1.triggered.connect(self.refresh_btn)  # Кнопка обновления данных в таблице
        self.delite_str.triggered.connect(self.delite_str_btn)  # Кнопка удаления
        self.exit_from_table.triggered.connect(self.exit_btn)  # Кнопка выхода
        self.MainWindow = ""

    def load_data_for_table(self):
        # Загрузка данных из БД в таблицу
        def set_row_in_table(row_from_db):
            # Метод для выравнивания значений в таблице по горизонтали и вертикали
            item = QtWidgets.QTableWidgetItem(row_from_db)
            item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
            return item

        # Подгрузка данных из дб
        data_plan = db.get_all_plan_for_every_mounth()
        self.maintable.setRowCount(len(data_plan))
        tablerow = 0
        for row in db.get_all_plan_for_every_mounth():
            col = 0
            for index in range(1, 5):
                self.maintable.setItem(tablerow, col, set_row_in_table(str(row[index])))
                col += 1
            tablerow += 1

    def add_str_btn(self):
        # Добавление данных в таблицу
        self.MainWindow = DialogAddPlanForMounth()
        self.MainWindow.show()

    def edit_str_btn(self):
        # Кнопка редактирования планового значения
        if self.maintable.selectedIndexes():
            row = self.maintable.currentIndex().row()
            for i in self.maintable.selectedIndexes():
                # получаем значение строки из 1 стобца(дата плана)
                date = self.maintable.model().index(row, 0).data()  # get cell at row, col
                kmd_plan = self.maintable.model().index(row, 1).data()
                tent_plan = self.maintable.model().index(row, 2).data()
                spu_plan = self.maintable.model().index(row, 3).data()
                data = [date, kmd_plan, tent_plan, spu_plan]
                self.MainWindow = DialogPlanEdit(data)
                self.MainWindow.show()
        else:
            row = self.maintable.rowCount() - 1

    def refresh_btn(self):
        # Кнопка обновления данных в таблице
        self.load_data_for_table()

    def delite_str_btn(self):
        # Удаление позиции планового значения
        date_for_plan = ''
        if self.maintable.selectedIndexes():
            row = self.maintable.currentIndex().row()
            date_for_plan = self.maintable.model().index(row, 0).data()  # get cell at row, col
        else:
            row = self.maintable.rowCount() - 1

        if row >= 0:
            title = "Подтверждение удаления позиции"
            text = f"Вы уверены, что хотите удалить данную позицию?\n"
            msg = MessageDialogWindow(title, text)  # Вызываем текстовое сообщение с подтверждением удаления
            if msg.conirm_message() == 1:
                self.maintable.removeRow(row)
                db.delite_plan(str(date_for_plan))
            else:
                self.update()

    def exit_btn(self):
        # Кнопка возвращения в главное меню
        self.close()
        self.MainWindow = Main_menu()
        self.MainWindow.show()


class DialogAddPlanForMounth(QDialog):
    # Добавление плана
    def __init__(self, parent=None, flag=Qt.Dialog):
        super().__init__(parent, flag)
        uic.loadUi('ui/db actions/add_plan_dialog.ui', self)
        self.dateEdit.setDate(QtCore.QDate.currentDate())  # Устанавливаем дату - сегодня
        self.dateEdit.setDisplayFormat('MMMM.yyyy')  # Устанавливаем формат даты
        self.add_btn_2.clicked.connect(self.add_prod_btn)  # Кнопка добавить
        self.MainWindow = ""

    def add_prod_btn(self):
        # Кнопка добавить
        # Получаем всю информацию из виджетов в окне
        date_plan = self.dateEdit.date().toString('MMMM yyyy')
        kmd_plan = self.kmd_SpinBox.value()
        tent_plan = self.tent_SpinBox.value()
        spu_plan = self.spu_SpinBox.value()
        date_real = self.dateEdit.date().toString('yyyy-MM-dd')
        try:
            db.add_plan_for_mounth(date_plan, kmd_plan, tent_plan, spu_plan, date_real)  # Добавляем инфу в БД
            # Выводим сообщение об успехе
            message_title = "Успешное добавление"
            message_text = f"Проект успешно добавлен в базу данных."
            msg = MessageDialogWindow(message_title, message_text)
            if msg.success_msg() == 1:
                self.close()
        except:
            error = 'Что-то пошло не так. Попробуйте снова ;).'
            self.MainWindow = ErrorAddReport(error)
            self.MainWindow.show()


class DialogPlanEdit(QDialog):
    # Редактирование планового значения
    def __init__(self, data, parent=None, flag=Qt.Dialog):
        super().__init__(parent, flag)
        uic.loadUi('ui/db actions/edit_plan_dialog.ui', self)
        self.change_btn.clicked.connect(self.change_prog_btn)  # Подключаем кнопку к эвенту
        self.data = data
        # Распредлеяем сохраненные данные из ячеек таблицы по виджетам
        for row in db.get_real_data_for_plan(self.data[0]):
            self.dateEdit.setDate(datetime.strptime(row[4], "%Y-%m-%d"))
        self.dateEdit.setDisplayFormat('MMMM.yyyy')
        self.kmd_SpinBox.setValue(float(self.data[1]))
        self.tent_SpinBox.setValue(float(self.data[2]))
        self.spu_SpinBox.setValue(float(self.data[3]))
        self.MainWindow = ""

    def change_prog_btn(self):
        try:
            date_real = self.dateEdit.date().toString('yyyy-MM-dd')
            kmd_plan = self.kmd_SpinBox.value()
            tent_plan = self.tent_SpinBox.value()
            spu_plan = self.spu_SpinBox.value()
            db.update_plan_for_mounth(kmd_plan, tent_plan, spu_plan, date_real)
            message_title = "Успешное изменение"
            message_text = f"Плановые значения успешно изменены."
            msg = MessageDialogWindow(message_title, message_text)
            if msg.success_msg() == 1:
                self.close()
        except:
            self.MainWindow = ErrorAddWin()
            self.MainWindow.show()


class ChooseProgect(QMainWindow):
    # Выбор таблицы данных
    def __init__(self, parent=None, flag=Qt.Window):
        super().__init__(parent, flag)
        # В зависиомсти от выбранных в главном меню настроек, загружаем окно выбра таблицы данных
        uic.loadUi('ui/choose_data.ui', self)
        self.option_menu = int(re.search('\d+', str(db.get_menu_main())).group(0))
        if self.option_menu == 2:
            self.print_tableButton.clicked.connect(self.data_table_push)
            self.backButton.clicked.connect(self.back_btn_push)
            self.progect_for_table = ''
            self.type_progBox.currentIndexChanged.connect(self.indexChanged)
            self.indexChanged(self.progectBox.currentIndex())
            self.progectBox_2.hide()

        elif self.option_menu == 3:
            self.type_progBox.hide()
            self.label_3.clear()
            self.print_tableButton.clicked.connect(self.data_table_montage_push)
            self.backButton.clicked.connect(self.back_btn_push)
            self.progect_for_table = ''
            self.progectbox_change_items()
        self.MainWindow = ""

    def indexChanged(self, index):
        self.progectBox.clear()
        data = []
        if self.type_progBox.currentText() == 'КМД':
            for kmd_prog in db.get_info_progect_kmd():
                data.append(kmd_prog)
        elif self.type_progBox.currentText() == 'СПУ':
            for spu_prog in db.get_info_progect_spu():
                data.append(spu_prog)
        elif self.type_progBox.currentText() == 'ПВХ':
            for pvh_prog in db.get_info_progect_pvh():
                data.append(pvh_prog)

        if data is not None:
            for prog in data:
                self.progectBox.addItems(prog)
        data.clear()

    def progectbox_change_items(self):
        self.progectBox.clear()
        progect_list = []
        for montage_prog in db.get_info_progect_montage():
            progect_list.append(montage_prog)
        for p in progect_list:
            self.progectBox_2.addItems(p)
        progect_list.clear()

    def back_btn_push(self):
        # Кнопка возвращения в главное меню
        self.close()
        self.MainWindow = Main_menu()
        self.MainWindow.show()

    def data_table_push(self):
        # Включение необходимой таблицы с данными по проекту производства
        self.progect_for_table = self.progectBox.currentText()
        type_document = self.type_progBox.currentText()
        if type_document == 'КМД':
            self.close()
            self.MainWindow = DataProgectTableCMD(self.progect_for_table)
            self.MainWindow.show()
        elif type_document == 'СПУ':
            self.close()
            self.MainWindow = DataProgectTableSPU(self.progect_for_table)
            self.MainWindow.show()
        elif type_document == 'ПВХ':
            self.close()
            self.MainWindow = DataProgectTablePVH(self.progect_for_table)
            self.MainWindow.show()

    def data_table_montage_push(self):
        # Включение необходимой таблицы с данными по проекту монтажа
        self.progect_for_table = self.progectBox_2.currentText()
        self.close()
        self.MainWindow = MontageTodayDataTable(self.progect_for_table)
        self.MainWindow.show()


class DataProgectTableCMD(QMainWindow):
    # КМД таблица данных
    def __init__(self, progect, parent=None, flag=Qt.Window):
        super().__init__(parent, flag)
        uic.loadUi('ui/table_with_data.ui', self)
        self.progect_name_for_table = progect
        self.progect_name_label.clear()
        self.progect_name_label.setText(
            f'Таблица данных (КМД) по {self.progect_name_for_table}')  # Установим виджет согласна проекта
        # Установим размеры столбцов
        self.maintable.setColumnWidth(0, 130)
        for num_of_col in range(1, 4):
            self.maintable.setColumnWidth(num_of_col, 100)
        self.maintable.setColumnWidth(4, 140)
        self.load_data_for_table()  # Загрузка данных в таблицу
        self.maintable.setWordWrap(True)
        self.refresh_1.triggered.connect(self.refresh_btn)  # Кнопка обновить
        self.delite_str.triggered.connect(self.delite_str_btn)  # Кнопка удалить
        self.exit_from_table.triggered.connect(self.exit_btn)  # Кнопка выхода
        self.MainWindow = ""

    def load_data_for_table(self):
        # Загрузка данных в таблицу
        data_prod = db.get_info_ab_day_report(self.progect_name_for_table)
        self.maintable.setRowCount(len(data_prod))
        tablerow = 0
        for row in db.get_info_ab_day_report(self.progect_name_for_table):
            col = 0
            for index in range(2, 7):
                self.maintable.setItem(tablerow, col, QtWidgets.QTableWidgetItem(str(row[index])))
                col += 1
            tablerow += 1

    def refresh_btn(self):
        # Кнопка обновления
        self.load_data_for_table()

    def delite_str_btn(self):
        # Кнопка удаления
        date = ''
        if self.maintable.selectedIndexes():
            row = self.maintable.currentIndex().row()
            # получаем значение строки из 1 стобца(наименование продукции)
            date = self.maintable.model().index(row, 0).data()  # get cell at row, col
        else:
            row = self.maintable.rowCount() - 1

        if row >= 0:
            title = "Подтверждение удаления позиции"
            text = f"Вы уверены, что хотите удалить данную позицию?\n"
            msg = MessageDialogWindow(title, text)  # Вызываем текстовое сообщение с подтверждением удаления
            if msg.conirm_message() == 1:
                self.maintable.removeRow(row)
                db.delit_data_report(self.progect_name_for_table, date)
            else:
                self.update()

    def exit_btn(self):
        # Кнопка выхода
        self.close()
        self.MainWindow = Main_menu()
        self.MainWindow.show()


class DataProgectTableSPU(QMainWindow):
    # СПУ таблица данных
    def __init__(self, progect, parent=None, flag=Qt.Window):
        super().__init__(parent, flag)
        uic.loadUi('ui/table_with_data_spu.ui', self)
        self.progect_name_for_table = progect
        self.progect_name_label.clear()
        # Установим название лейбла
        self.progect_name_label.setText(f'Таблица данных (СПУ) по {self.progect_name_for_table}')
        # Установим размеры столбцов
        self.maintable.setColumnWidth(0, 110)
        for colm_num in range(1, 7):
            if colm_num == 4 or colm_num == 5 or colm_num == 6:
                self.maintable.setColumnWidth(colm_num, 150)
            else:
                self.maintable.setColumnWidth(colm_num, 135)
            colm_num += 1
        self.load_data_for_table()  # Загрузка данных в таблицу
        self.maintable.setWordWrap(True)
        self.refresh_1.triggered.connect(self.refresh_btn)  # Кнопка обновления данных в таблице
        self.delite_str.triggered.connect(self.delite_str_btn)  # Кнопка удаления
        self.exit_from_table.triggered.connect(self.exit_btn)  # Кнопка выхода
        self.MainWindow = ""

    def load_data_for_table(self):
        # Загрузка данных в таблицу
        # Получаем данные из БД
        data_prod = db.get_info_ab_day_report_spu(self.progect_name_for_table)
        self.maintable.setRowCount(len(data_prod))  # Устнавливаем кол-во стобцов
        tablerow = 0
        for row in db.get_info_ab_day_report_spu(self.progect_name_for_table):
            col = 0
            for index in range(2, 9):
                self.maintable.setItem(tablerow, col, QtWidgets.QTableWidgetItem(str(row[index])))
                col += 1
            tablerow += 1

    def refresh_btn(self):
        # Кнопка обновления данных в таблице
        self.load_data_for_table()

    def delite_str_btn(self):
        # Кнопка удаления
        date = ''
        if self.maintable.selectedIndexes():
            row = self.maintable.currentIndex().row()
            date = self.maintable.model().index(row, 0).data()  # get cell at row, col
        else:
            row = self.maintable.rowCount() - 1

        if row >= 0:
            title = "Подтверждение удаления позиции"
            text = f"Вы уверены, что хотите удалить данную позицию?\n"
            msg = MessageDialogWindow(title, text)  # Вызываем текстовое сообщение с подтверждением удаления
            if msg.conirm_message() == 1:
                self.maintable.removeRow(row)
                db.delit_report_spu_data(self.progect_name_for_table, date)
            else:
                self.update()

    def exit_btn(self):
        # Кнопка выхода
        self.close()
        self.MainWindow = Main_menu()
        self.MainWindow.show()


class DataProgectTablePVH(QMainWindow):
    # ПВХ таблица данных
    def __init__(self, progect, parent=None, flag=Qt.Window):
        super().__init__(parent, flag)
        uic.loadUi('ui/table_with_data_pvh.ui', self)
        self.progect_name_for_table = progect
        self.progect_name_label.clear()
        self.progect_name_label.setText(f'Таблица данных (ПВХ) по {self.progect_name_for_table}')
        # Установим размеры столбцов
        self.maintable.setColumnWidth(0, 110)
        for colm_num in range(1, 9):
            if colm_num == 2 or colm_num == 4:
                self.maintable.setColumnWidth(colm_num, 140)
            elif colm_num == 8:
                self.maintable.setColumnWidth(colm_num, 150)
            else:
                self.maintable.setColumnWidth(colm_num, 135)

        self.load_data_for_table()  # Загрузка данных в таблицу
        self.maintable.setWordWrap(True)
        self.refresh_1.triggered.connect(self.refresh_btn)  # Кнопка обновления данных в таблице
        self.delite_str.triggered.connect(self.delite_str_btn)  # Кнопка удаления
        self.exit_from_table.triggered.connect(self.exit_btn)  # Кнопка выхода
        self.MainWindow = ""

    def load_data_for_table(self):
        # Загрузка данных в таблицу
        data_prod = db.get_info_ab_day_report_pvh(self.progect_name_for_table)
        self.maintable.setRowCount(len(data_prod))
        tablerow = 0
        for row in db.get_info_ab_day_report_pvh(self.progect_name_for_table):
            col = 0
            for index in range(2, 11):
                self.maintable.setItem(tablerow, col, QtWidgets.QTableWidgetItem(str(row[index])))
                col += 1
            tablerow += 1

    def refresh_btn(self):
        # Кнопка обновления данных в таблице
        self.load_data_for_table()

    def delite_str_btn(self):
        # Кнопка удаления
        date = ''
        if self.maintable.selectedIndexes():
            row = self.maintable.currentIndex().row()
            for i in self.maintable.selectedIndexes():
                # получаем значение строки из 1 стобца(наименование продукции)
                date = self.maintable.model().index(row, 0).data()  # get cell at row, col
        else:
            row = self.maintable.rowCount() - 1

        if row >= 0:
            title = "Подтверждение удаления позиции"
            text = f"Вы уверены, что хотите удалить данную позицию?\n"
            msg = MessageDialogWindow(title, text)  # Вызываем текстовое сообщение с подтверждением удаления
            if msg.conirm_message() == 1:
                self.maintable.removeRow(row)
                db.delit_report_pvh_data(self.progect_name_for_table, date)
            else:
                self.update()

    def exit_btn(self):
        # Кнопка выхода
        self.close()
        self.MainWindow = Main_menu()
        self.MainWindow.show()


class MontageProgTable(QMainWindow):
    # Табица проектов
    def __init__(self, parent=None, flag=Qt.Window):
        super().__init__(parent, flag)
        uic.loadUi('ui/montage_table/montage_progect_table.ui', self)
        # Установим ширину столбцов
        self.maintable.setColumnWidth(0, 300)
        self.maintable.setColumnWidth(1, 200)
        self.load_data_for_table()  # Подгрузка данных из дб
        self.maintable.setWordWrap(True)
        self.add_str.triggered.connect(self.add_str_btn)  # Кнопка добавления данных в таблицу и БД
        self.edit_str.triggered.connect(self.edit_str_btn)  # Кнопка редакт. данных в таблице и БД
        self.refresh_1.triggered.connect(self.refresh_btn)  # Кнопка обновления данных в таблице
        self.delite_str.triggered.connect(self.delite_str_btn)  # Кнопка удаления
        self.exit_from_table.triggered.connect(self.exit_btn)  # Кнопка выхода
        self.MainWindow = ""

    def load_data_for_table(self):
        def set_row_in_table(row_from_db):
            # Метод для выравнивания значений в таблице по горизонтали и вертикали
            item = QtWidgets.QTableWidgetItem(row_from_db)
            item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
            return item

        # Подгрузка данных из дб
        data_prod = db.get_all_info_progect()
        self.maintable.setRowCount(len(data_prod))
        tablerow = 0
        for row in db.get_all_info_progect():
            self.maintable.setItem(tablerow, 0, QtWidgets.QTableWidgetItem(row[1]))
            if row[6] is True or row[6] == 1:
                self.maintable.setItem(tablerow, 1, set_row_in_table('V'))
            else:
                self.maintable.setItem(tablerow, 1, set_row_in_table('X'))
            tablerow += 1

    def add_str_btn(self):
        # Кнопка добавления данных в таблицу и БД
        self.MainWindow = MontageDialogAdd()
        self.MainWindow.show()

    def edit_str_btn(self):
        # Кнопка редакт. данных в таблице и БД
        if self.maintable.selectedIndexes():
            row = self.maintable.currentIndex().row()
            # получаем значение строки из 1 стобца(наименование продукции)
            data = [self.maintable.model().index(row, 0).data()]  # get cell at row, col
            self.MainWindow = MontageDialogEdit(data)
            self.MainWindow.show()
        else:
            row = self.maintable.rowCount() - 1

    def refresh_btn(self):
        # Кнопка обновления данных в таблице
        self.load_data_for_table()

    def delite_str_btn(self):
        # Удаление данных
        progect = ''
        if self.maintable.selectedIndexes():
            row = self.maintable.currentIndex().row()
            # получаем значение строки из 1 стобца(наименование продукции)
            progect = self.maintable.model().index(row, 0).data()  # get cell at row, col
        else:
            row = self.maintable.rowCount() - 1

        if row >= 0:
            title = "Подтверждение удаления позиции"
            text = f"Вы уверены, что хотите удалить данную позицию?\n"
            msg = MessageDialogWindow(title, text)  # Вызываем текстовое сообщение с подтверждением удаления
            if msg.conirm_message() == 1:
                self.maintable.removeRow(row)
                db.delite_prod(str(progect))
            else:
                self.update()

    def exit_btn(self):
        # Кнопка выхода
        self.close()
        self.MainWindow = Main_menu()
        self.MainWindow.show()


class MontageDialogAdd(QDialog):
    # Добавление проекта в таблицу и БД
    def __init__(self, parent=None, flag=Qt.Dialog):
        super().__init__(parent, flag)
        uic.loadUi('ui/montage_table/add_montage_dialog.ui', self)
        self.add_btn_2.clicked.connect(self.add_prod_btn)

    def add_prod_btn(self):
        # Кнопка Добавление проекта в таблицу и БД
        prog = self.textEdit_prog.toPlainText()  # Берем текст из виджета
        db.add_progect_name(prog)  # Добавляем в базу данных
        # Обновляем статус проекта в базе данных, в случае если установлен флаг в чекбоксе
        if self.montage_checkBox.isChecked():
            db.update_active_montage(prog)
        else:
            db.update_disactive_montage(prog)
        # Выводим сообщение об успехе
        message_title = "Успешное добавление"
        message_text = f"Проект успешно добавлен в базу данных."
        msg = MessageDialogWindow(message_title, message_text)
        if msg.success_msg() == 1:
            self.close()


class MontageDialogEdit(QDialog):
    # Редактирование позиции в таблице
    def __init__(self, data, parent=None, flag=Qt.Dialog):
        super().__init__(parent, flag)
        uic.loadUi('ui/montage_table/edit_montage_dialog.ui', self)
        self.change_btn.clicked.connect(self.change_prog_btn)
        self.data = data
        progect = ''.join(self.data[0])  # Получаем название проекта
        self.textEdit_prog.setText(f"{progect}")  # Уставливаем название проекта в виджет
        self.MainWindow = ""
        self.id = str(int(re.search('\d+', str(db.get_id_progect(progect))).group(0)))  # Ищем id проекта
        for p in db.get_info_about_this_progect(progect):
            if p[6] is True or p[6] == 1:
                self.montage_checkBox.setChecked(True)

    def change_prog_btn(self):
        # Кнопка подтверждения изменения
        try:
            progect = self.textEdit_prog.toPlainText()
            db.update_progect_name(progect, int(self.id))
            if self.montage_checkBox.isChecked():
                db.update_active_montage(progect)
            else:
                db.update_disactive_montage(progect)
            # Выводим сообщение об успехе
            message_title = "Успешное изменение"
            message_text = f"Плановые значения успешно изменены."
            msg = MessageDialogWindow(message_title, message_text)
            if msg.success_msg() == 1:
                self.close()
        except:
            self.MainWindow = ErrorAddWin()
            self.MainWindow.show()


class MontageTodayDataTable(QMainWindow):
    # Значения по монтажы ежедневные по проекту
    def __init__(self, progect, parent=None, flag=Qt.Window):
        super().__init__(parent, flag)
        uic.loadUi('ui/table_with_data_montage.ui', self)
        self.progect_name_for_table = progect
        self.progect_name_label.clear()
        self.progect_name_label.setText(f'Таблица данных (Монтаж) по {self.progect_name_for_table}')
        # Установка рамзера ширины колонок
        for colm_num in range(1, 9):
            if colm_num == 2 or colm_num == 4:
                self.maintable.setColumnWidth(colm_num, 140)
            elif colm_num == 6 or colm_num == 8:
                self.maintable.setColumnWidth(colm_num, 150)
            else:
                self.maintable.setColumnWidth(colm_num, 135)
        self.load_data_for_table()  # Загрузка данных из БД
        self.maintable.setWordWrap(True)
        self.refresh_1.triggered.connect(self.refresh_btn)  # Обновление данных в таблице
        self.delite_str.triggered.connect(self.delite_str_btn)  # Удаление данных
        self.exit_from_table.triggered.connect(self.exit_btn)  # Выход
        self.MainWindow = ""

    def load_data_for_table(self):
        # Загрузка данных из БД
        data_prod = db.get_info_evr_report_montage(self.progect_name_for_table)
        self.maintable.setRowCount(len(data_prod))
        tablerow = 0
        for row in db.get_info_evr_report_montage(self.progect_name_for_table):
            col = 0
            for index in range(2, 11):
                self.maintable.setItem(tablerow, col, QtWidgets.QTableWidgetItem(str(row[index])))
                col += 1
            tablerow += 1

    def refresh_btn(self):
        # Обновление данных в таблице
        self.load_data_for_table()

    def delite_str_btn(self):
        # Удаление данных
        date = ''
        if self.maintable.selectedIndexes():
            row = self.maintable.currentIndex().row()
            date = self.maintable.model().index(row, 0).data()  # get cell at row, col
        else:
            row = self.maintable.rowCount() - 1

        if row >= 0:
            title = "Подтверждение удаления позиции"
            text = f"Вы уверены, что хотите удалить данную позицию?\n"
            msg = MessageDialogWindow(title, text)  # Вызываем текстовое сообщение с подтверждением удаления
            if msg.conirm_message() == 1:
                self.maintable.removeRow(row)
                db.ddelite_data_montag_report_by_date(self.progect_name_for_table, date)
            else:
                self.update()

    def exit_btn(self):
        self.close()
        self.MainWindow = Main_menu()
        self.MainWindow.show()


class ErrorAddWin(QDialog):
    # Диалоговое окно ошибки добавление в базу данных
    def __init__(self, parent=None, flag=Qt.Dialog):
        super().__init__(parent, flag)
        uic.loadUi('ui/errors/error_dialog.ui', self)
        self.ok_btn.clicked.connect(self.back_btn_push)

    def back_btn_push(self):
        self.close()


class AddData(QMainWindow):
    # Добавление данных из excel в БД
    def __init__(self, parent=None, flag=Qt.Window):
        super().__init__(parent, flag)
        uic.loadUi('ui/add_data.ui', self)
        self.backButton.clicked.connect(self.back_btn_push)  # Кнопка возвращения в главное меню
        self.chooseButton.clicked.connect(self.choose_btn_push)  # Кнопка выбора excel file
        self.add_data_from_excelButton.clicked.connect(self.add_data_in_bd_push)  # Кнопка добалвения данных в excel
        self.reportButton.clicked.connect(self.report_bd_push)  # Кнопка вывода отчета в excel
        self.dateEdit.setDate(QtCore.QDate.currentDate())  # Устанавливаем текущую дату в виджет
        self.dateEdit.setDisplayFormat('dd.MM.yyyy')  # Формат вывода даты в виджете
        # Назначаем переменные в конструкторе, для дальнейшего использования по классу
        self.filename = ''
        self.filetype = ''
        self.path_to_exel = ""
        self.data_perc = []
        self.progect_squer = 0
        self.type_progBox.currentIndexChanged.connect(self.indexChanged)  # Управление комбобоксом
        self.indexChanged()  # Управление комбобоксом
        self.MainWindow = ""

    def indexChanged(self):
        # Управление комбобоксом
        self.progectBox.clear()  # Очищаем комбобокс от данных
        data = []
        # Добавляем активные проекты из БД в зависимости от КД
        if self.type_progBox.currentText() == 'КМД':
            for kmd_prog in db.get_info_progect_kmd():
                data.append(kmd_prog)
        elif self.type_progBox.currentText() == 'СПУ':
            for spu_prog in db.get_info_progect_spu():
                data.append(spu_prog)
        elif self.type_progBox.currentText() == 'ПВХ':
            for pvh_prog in db.get_info_progect_pvh():
                data.append(pvh_prog)
        # Добавляем полученный результат в комбобокс
        if data is not None:
            for prog in data:
                self.progectBox.addItems(prog)
        data.clear()

    def back_btn_push(self):
        # Кнопка возвращения в главное меню
        self.close()
        self.MainWindow = Main_menu()
        self.MainWindow.show()

    def choose_btn_push(self):
        # Кнопка выбора excel file
        def calculate_values(list_of_data, sum):
            # Вычисляем суммарный процент готовности и сумарную площадь\массу продукции
            common_ready_perc = 0
            common_sum_ready = 0
            for evr_data in list_of_data:
                per_massa = round((evr_data / sum * 100), 2)
                common_ready_perc += per_massa
                common_sum_ready += evr_data
            ready_data = [common_ready_perc, common_sum_ready]
            return ready_data

        self.data_perc.clear()
        self.filename, self.filetype = QFileDialog.getOpenFileName(self,
                                                                   "Выбрать файл",
                                                                   ".",
                                                                   "Excel Files(*.xlsx);;All Files(*)")
        self.plainTextEdit.appendHtml("Путь: <b>{}</b> <br> <b>{:*^54}</b>"
                                      "".format(self.filename, self.filetype))
        self.path_to_exel = "{}".format(self.filename)
        try:
            wb = load_workbook(f"{self.path_to_exel}", data_only=True)
            if 'еталл' in self.filename or 'Металл' in self.filename or 'металл' \
                    in self.filename or 'КМД' in self.filename:
                # Значение массы по проекту Заготовка
                blank_massa = 0
                sawing_corners_massa = 0
                drill_massa = 0
                plasma_massa = 0
                # Значение массы по факту на тек. момент Заготовка
                bl_ready_massa = 0
                sc_ready_massa = 0
                d_ready_massa = 0
                p_ready_massa = 0
                # Значение массы по проекту Сварка
                weld_assambl_massa = 0
                weld_massa = 0
                # Значение массы по факту на тек. момент Сварка
                wa_ready_massa = 0
                w_ready_massa = 0
                # Остальное
                painting_massa = 0
                paint_ready_massa = 0
                try:
                    sheet_ranges_1 = wb[f'Заготовка']
                    blank_massa = round(sheet_ranges_1['G3'].value, 2)
                    bl_ready_massa = round(sheet_ranges_1['M3'].value, 2)
                except:
                    error = 'Отсутсвует значение в ячейке G3 или М3 (Заготовка).' \
                            ' Исправьте файл и попробуйте снова.'
                    self.MainWindow = ErrorAddReport(error)
                    self.MainWindow.show()
                try:
                    sheet_ranges_2 = wb[f'Распиловка углов']
                    sawing_corners_massa = round(sheet_ranges_2['G3'].value, 2)
                    sc_ready_massa = round(sheet_ranges_2['M3'].value, 2)
                except:
                    error = 'Отсутсвует значение в ячейке G3 или М3 (Распил. углов).' \
                            ' Исправьте файл и попробуйте снова.'
                    self.MainWindow = ErrorAddReport(error)
                    self.MainWindow.show()

                try:
                    sheet_ranges_3 = wb[f'Сверлильный станок']
                    drill_massa = round(sheet_ranges_3['G3'].value, 2)
                    d_ready_massa = round(sheet_ranges_3['M3'].value, 2)
                except:
                    error = 'Отсутсвует значение в ячейке G3 или М3 (Сверл. станок).' \
                            ' Исправьте файл и попробуйте снова.'
                    self.MainWindow = ErrorAddReport(error)
                    self.MainWindow.show()

                try:
                    sheet_ranges_4 = wb[f'Участок резки плазмой']
                    plasma_massa = round(sheet_ranges_4['G3'].value, 2)
                    p_ready_massa = round(sheet_ranges_4['M3'].value, 2)
                except:
                    error = 'Отсутсвует значение в ячейке G3 или М3 (Плазмен. станок).' \
                            ' Исправьте файл и попробуйте снова.'
                    self.MainWindow = ErrorAddReport(error)
                    self.MainWindow.show()

                massa_ready_blank = [bl_ready_massa, sc_ready_massa, d_ready_massa, p_ready_massa]
                common_blank_massa_from_prog = blank_massa + sawing_corners_massa + drill_massa + plasma_massa
                calculate_blank_massa = calculate_values(massa_ready_blank, common_blank_massa_from_prog)
                self.data_perc.append(round(calculate_blank_massa[0], 2))
                try:
                    sheet_ranges_5 = wb[f'Сборка']
                    weld_assambl_massa = round(sheet_ranges_5['F3'].value, 2)
                    wa_ready_massa = round(sheet_ranges_5['L3'].value, 2)
                except:
                    error = 'Отсутсвует значение в ячейке F3 или L3 (Сборка). Исправьте файл и попробуйте снова.'
                    self.MainWindow = ErrorAddReport(error)
                    self.MainWindow.show()

                try:
                    sheet_ranges_6 = wb[f'Участок сварки']
                    weld_massa = round(sheet_ranges_6['F3'].value, 2)
                    w_ready_massa = round(sheet_ranges_6['L3'].value, 2)
                except:
                    error = 'Отсутсвует значение в ячейке F3 или L3 (Участок сварки).' \
                            ' Исправьте файл и попробуйте снова.'
                    self.MainWindow = ErrorAddReport(error)
                    self.MainWindow.show()
                weld_ready_data = [wa_ready_massa, w_ready_massa]
                sum_weld_massa_from_prog = weld_assambl_massa + weld_massa
                # Вычисляем общий процент готовности и сумарную массу сварки
                calculate_welding = calculate_values(weld_ready_data, sum_weld_massa_from_prog)
                self.data_perc.append(round(calculate_welding[0], 2))
                try:
                    sheet_ranges_7 = wb[f'Покраска']
                    painting_massa = round(sheet_ranges_7['F3'].value, 2)
                    paint_ready_massa = round(sheet_ranges_7['L3'].value, 2)
                    paint_perc = round(sheet_ranges_7['J3'].value * 100, 2)
                    self.data_perc.append(paint_perc)
                except:
                    error = 'Отсутсвует значение в ячейке F3 или L3 (Покраска).' \
                            ' Исправьте файл и попробуйте снова.'
                    self.MainWindow = ErrorAddReport(error)
                    self.MainWindow.show()
                common_ready_massa_data = [calculate_blank_massa[1], calculate_welding[1], paint_ready_massa]
                summa_massa_for_prog = common_blank_massa_from_prog + sum_weld_massa_from_prog + painting_massa
                # Вычисляем общий процент готовности и сумарную массу КМД
                calculate_data = calculate_values(common_ready_massa_data, summa_massa_for_prog)
                self.data_perc.append(calculate_data[0])
                self.data_perc.append(calculate_data[1])
                self.data_perc.append(summa_massa_for_prog)
                print(common_ready_massa_data)

            elif 'СПУ' in self.filename or 'утеплит' in self.filename \
                    or 'утеплитель' in self.filename or 'Утеплитель' in self.filename or 'теплит' in self.filename:
                # Значение площади по проекту
                rpolym_squer = 0
                sew_loops_squer = 0
                stit_spu_squer = 0
                gloe_p_squer = 0
                assembl_p_squer = 0
                punch_l_squer = 0
                # Значение площади изготовленной продукции на тек. момент
                rpolym_ready_squer = 0
                sew_loops_ready_squer = 0
                stit_spu_ready_squer = 0
                gloe_p_ready_squer = 0
                assembl_p_ready_squer = 0
                punch_l_ready_squer = 0
                try:
                    sheet_1 = wb[f"Раскрой полипропилена"]
                    raskroy_polypr = round(sheet_1['K3'].value * 100, 2)
                    rpolym_squer = round(sheet_1['G3'].value, 2)
                    rpolym_ready_squer = round(sheet_1['M3'].value, 2)
                    self.data_perc.append(raskroy_polypr)
                except:
                    error = 'Отсутсвует значение в ячейках K3, G3, M3' \
                            ' (Раскрой полипроп.). Исправьте файл.'
                    self.MainWindow = ErrorAddReport(error)
                    self.MainWindow.show()

                try:
                    sheet_2 = wb[f'Пришить крючки и петли']
                    sew_loops = round(sheet_2['K3'].value * 100, 2)
                    sew_loops_squer = round(sheet_2['G3'].value, 2)
                    sew_loops_ready_squer = round(sheet_2['M3'].value, 2)
                    self.data_perc.append(sew_loops)
                except:
                    error = 'Отсутсвует значение в ячейках K3, G3, M3 ' \
                            '(Пришить крючки и петли). Исправьте файл.'
                    self.MainWindow = ErrorAddReport(error)
                    self.MainWindow.show()

                try:
                    sheet_3 = wb[f'Сшивка полипропилена']
                    stitching = round(sheet_3['K3'].value * 100, 2)
                    stit_spu_squer = round(sheet_3['G3'].value, 2)
                    stit_spu_ready_squer = round(sheet_3['M3'].value, 2)
                    self.data_perc.append(stitching)
                except:
                    error = 'Отсутсвует значение в ячейках K3, G3, M3' \
                            ' (Сшивка полипропилена). Исправьте файл.'
                    self.MainWindow = ErrorAddReport(error)
                    self.MainWindow.show()

                try:
                    sheet_4 = wb[f'Наклейка синтепона']
                    gloe_polypr = round(sheet_4['J3'].value * 100, 2)
                    gloe_p_squer = round(sheet_4['F3'].value, 2)
                    gloe_p_ready_squer = round(sheet_4['L3'].value, 2)
                    self.data_perc.append(gloe_polypr)
                except:
                    error = 'Отсутсвует значение в ячейках J3, F3, L3' \
                            ' (Наклейка синтепона). Исправьте файл.'
                    self.MainWindow = ErrorAddReport(error)
                    self.MainWindow.show()

                try:
                    sheet_5 = wb[f'Сборка']
                    assembl_polypr = round(sheet_5['J3'].value * 100, 2)
                    assembl_p_squer = round(sheet_5['F3'].value, 2)
                    assembl_p_ready_squer = round(sheet_5['L3'].value, 2)
                    self.data_perc.append(assembl_polypr)
                except:
                    error = 'Отсутсвует значение в ячейках J3, F3, L3 (Сборка). Исправьте файл.'
                    self.MainWindow = ErrorAddReport(error)
                    self.MainWindow.show()

                try:
                    sheet_6 = wb[f'Пробивка люверс']
                    punch_luverc = round(sheet_6['J3'].value * 100, 2)
                    punch_l_squer = round(sheet_6['F3'].value, 2)
                    punch_l_ready_squer = round(sheet_6['L3'].value, 2)
                    self.data_perc.append(punch_luverc)
                except:
                    error = 'Отсутсвует значение в ячейках J3, F3, L3' \
                            ' (Пробивка люверс). Исправьте файл.'
                    self.MainWindow = ErrorAddReport(error)
                    self.MainWindow.show()

                try:
                    sheet_7 = wb[f'Упаковка']
                    ready_prod = round(sheet_7['J3'].value * 100, 2)
                    self.data_perc.append(ready_prod)
                except:
                    error = 'Отсутсвует значение в ячейке J3 (Упаковка). Исправьте файл.'
                    self.MainWindow = ErrorAddReport(error)
                    self.MainWindow.show()
                squer_data_spu = [rpolym_ready_squer, sew_loops_ready_squer, stit_spu_ready_squer,
                                  gloe_p_ready_squer, assembl_p_ready_squer, punch_l_ready_squer]
                common_summ_squer_from_prog = rpolym_squer + sew_loops_squer + stit_spu_squer \
                                              + gloe_p_squer + assembl_p_squer + punch_l_squer
                # Вычисляем общий процент готовности и сумарную площадь СПУ
                calculate_data = calculate_values(squer_data_spu, common_summ_squer_from_prog)
                self.data_perc.append(calculate_data[0])
                self.data_perc.append(common_summ_squer_from_prog)
                self.data_perc.append(calculate_data[1])

            elif 'ПВХ' in self.filename or 'пвх' in self.filename or 'тентовое' in self.filename \
                    or 'полотно' in self.filename or 'тент' in self.filename:
                # Значение площади продукции по проекту
                rpolt_squer = 0
                rpock_squer = 0
                nashel_squer = 0
                wp_squer = 0
                wip_squer = 0
                slit_squer = 0
                wn_squer = 0
                # Значение площади изготовленной продукции на тек. момент
                rpolt_ready_s = 0
                rpock_ready_s = 0
                nashel_ready_s = 0
                wp_ready_s = 0
                wip_ready_s = 0
                slit_ready_s = 0
                wn_ready_s = 0
                try:
                    sheet_1 = wb[f"Раскрой полотна"]
                    raskroy_polotna = round(sheet_1['K3'].value * 100, 2)
                    rpolt_squer = round(sheet_1['G3'].value, 2)
                    rpolt_ready_s = round(sheet_1['M3'].value, 2)
                    self.data_perc.append(raskroy_polotna)
                except:
                    error = 'Отсутсвует значение в ячейках К3, G3, M3(1 лист). Исправьте файл.'
                    self.MainWindow = ErrorAddReport(error)
                    self.MainWindow.show()
                try:
                    sheet_2 = wb[f'Раскрой карманов']
                    raskroy_pockets = round(sheet_2['K3'].value * 100, 2)
                    rpock_squer = round(sheet_2['G3'].value, 2)
                    rpock_ready_s = round(sheet_2['M3'].value, 2)
                    self.data_perc.append(raskroy_pockets)
                except:
                    error = 'Отсутсвует значение в ячейках К3, G3, M3(2 лист). Исправьте файл.'
                    self.MainWindow = ErrorAddReport(error)
                    self.MainWindow.show()
                try:
                    try:
                        sheet_3 = wb[f'Раскрой нащельников']
                        nashelnik = round(sheet_3['J3'].value * 100, 2)
                        nashel_squer = round(sheet_3['F3'].value, 2)
                        nashel_ready_s = round(sheet_3['L3'].value, 2)
                        self.data_perc.append(nashelnik)
                    except:
                        sheet_3 = wb[f'Нащельники']
                        nashelnik = round(sheet_3['J3'].value * 100, 2)
                        nashel_squer = round(sheet_3['F3'].value, 2)
                        nashel_ready_s = round(sheet_3['L3'].value, 2)
                        self.data_perc.append(nashelnik)
                except:
                    error = 'Отсутсвует значение значениев ячейках J3, F3, L3(Раскрой нащельников). Исправьте файл.'
                    self.MainWindow = ErrorAddReport(error)
                    self.MainWindow.show()
                try:
                    sheet_4 = wb[f'Сварка карманов']
                    weld_pockets = round(sheet_4['K3'].value * 100, 2)
                    wp_squer = round(sheet_4['G3'].value, 2)
                    wp_ready_s = round(sheet_4['M3'].value, 2)
                    self.data_perc.append(weld_pockets)
                except:
                    error = 'Отсутсвует значение в ячейках К3, G3, M3(Сварка карманов). Исправьте файл.'
                    self.MainWindow = ErrorAddReport(error)
                    self.MainWindow.show()
                try:
                    sheet_5 = wb[f'Приварить карманы']
                    weld_in_pockets = round(sheet_5['J3'].value * 100, 2)
                    wip_squer = round(sheet_5['F3'].value, 2)
                    wip_ready_s = round(sheet_5['L3'].value, 2)
                    self.data_perc.append(weld_in_pockets)
                except:
                    error = 'Отсутсвует значениев ячейках J3, F3, L3 (Приварить карманы). Исправьте файл.'
                    self.MainWindow = ErrorAddReport(error)
                    self.MainWindow.show()
                try:
                    sheet_6 = wb[f'Пришить полосу второго слоя']
                    stitching = round(sheet_6['J3'].value * 100, 2)
                    slit_squer = round(sheet_6['F3'].value, 2)
                    slit_ready_s = round(sheet_6['L3'].value, 2)
                    self.data_perc.append(stitching)
                except:
                    error = 'Отсутсвует значение в ячейках J3, F3, L3 (Пришить полосу ВС). Исправьте файл.'
                    self.MainWindow = ErrorAddReport(error)
                    self.MainWindow.show()
                try:
                    sheet_7 = wb[f'Приварить нащельники']
                    weld_nashelnik = round(sheet_7['J3'].value * 100, 2)
                    wn_squer = round(sheet_7['F3'].value * 100, 2)
                    wn_ready_s = round(sheet_7['L3'].value * 100, 2)
                    self.data_perc.append(weld_nashelnik)
                except:
                    error = 'Отсутсвует значение в ячейках J3, F3, L3 (Приварить нащельники). Исправьте файл.'
                    self.MainWindow = ErrorAddReport(error)
                    self.MainWindow.show()
                try:
                    sheet_8 = wb[f'Упаковка продукции']
                    ready_prod = round(sheet_8['J3'].value * 100, 2)
                    self.data_perc.append(ready_prod)
                except:
                    error = 'Отсутсвует значение в ячейках J3, F3, L3 (Упаковка продукции). Исправьте файл.'
                    self.MainWindow = ErrorAddReport(error)
                    self.MainWindow.show()
                squer_ready_data = [rpolt_ready_s, rpock_ready_s, nashel_ready_s, wp_ready_s, wip_ready_s, slit_ready_s,
                                    wn_ready_s]
                common_sqeuer_of_prog = rpolt_squer + rpock_squer + nashel_squer + wp_squer + wip_squer + slit_squer + wn_squer
                # Вычисляем общий процент готовности и сумарную площадь ПВХ
                calculate_data = calculate_values(squer_ready_data, common_sqeuer_of_prog)
                self.data_perc.append(calculate_data[0])
                self.data_perc.append(common_sqeuer_of_prog)
                self.data_perc.append(calculate_data[1])
            self.success_2.clear()
        except:
            print(self.filename)
            error = 'Данные из файла не выбраны. Проверьте файл и попробуйте снова.'
            self.MainWindow = ErrorAddReport(error)
            self.MainWindow.show()

    def add_data_in_bd_push(self):
        # Добавление данных в БД
        try:
            if self.data_perc:
                progect_name = self.progectBox.currentText()  # Берем данны из текстового виджета
                date = self.dateEdit.date().toString('dd MMM yy')  # Берем данные из виджетат с датой
                # Подбираем ключ относительно проекта
                key = str(int(re.search('\d+', str(db.get_id_progect(progect_name))).group(0)) + 1)
                # Берем еще один формат даты для БД
                real_date = self.dateEdit.date().toString('yyyy.MM.dd')
                # Добавляем данные относительно какой КД они принадлежат в БД
                if 'еталл' in self.filename or 'Металл' in self.filename \
                        or 'металл' in self.filename or 'КМД' in self.filename:
                    db.delit_data_report(progect_name, date)
                    db.add_reporting_data(progect_name, date, self.data_perc[0],
                                          self.data_perc[1], self.data_perc[2],
                                          self.data_perc[3], self.data_perc[4], self.data_perc[5], key, real_date)
                elif 'СПУ' in self.filename or 'утеплит' in self.filename \
                        or 'утеплитель' in self.filename or 'Утеплитель' in self.filename or 'теплит' in self.filename:
                    db.delit_report_spu_data(progect_name, date)
                    db.add_report_spu(progect_name, date, self.data_perc[0], self.data_perc[1], self.data_perc[2],
                                      self.data_perc[3], self.data_perc[4], self.data_perc[5], self.data_perc[6],
                                      self.data_perc[7], self.data_perc[8], self.data_perc[9], key, real_date)
                elif 'ПВХ' in self.filename or 'пвх' in self.filename or 'тентовое' in self.filename \
                        or 'полотно' in self.filename or 'тент' in self.filename:
                    db.delit_report_pvh_data(progect_name, date)
                    db.add_report_pvh(progect_name, date, self.data_perc[0], self.data_perc[1],
                                      self.data_perc[2], self.data_perc[3], self.data_perc[4],
                                      self.data_perc[5], self.data_perc[6], self.data_perc[7],
                                      self.data_perc[8], self.data_perc[9], self.data_perc[10], key, real_date)
                self.success_2.setPixmap(QPixmap("images/dop/success.png"))
                message_title = "Успешное портирование данных"
                message_text = f"Плановые значения успешно изменены."
                msg = MessageDialogWindow(message_title, message_text)
                if msg.success_msg() == 1:
                    pass
                self.data_perc.clear()
            else:
                error = 'Выберите файл, а затем добавьте данные.'
                self.MainWindow = ErrorAddReport(error)
                self.MainWindow.show()
        except:
            error = 'Ошибка добавления данных в базу данных. Попробуйте снова.'
            self.MainWindow = ErrorAddReport(error)
            self.MainWindow.show()

    def report_bd_push(self):
        try:
            type_progect = self.type_progBox.currentText()
            date = self.dateEdit.date().toString('dd MMM yy')
            date_for_plan = self.dateEdit.date().toString('MMMM yyyy')
            excel(type_progect, date, date_for_plan)
        except:
            error = 'Ошибка cоздания отчета. Возможно файл уже открыт. Попробуйте снова.'
            self.MainWindow = ErrorAddReport(error)
            self.MainWindow.show()


class AddMontageData(QMainWindow):
    # Добавление данных от монтажных подразделений в БД
    def __init__(self, parent=None, flag=Qt.Window):
        super().__init__(parent, flag)
        uic.loadUi('ui/add_montage_data.ui', self)
        self.backButton.clicked.connect(self.back_btn_push)
        self.chooseButton.clicked.connect(self.choose_report_btn_push)
        self.add_data_from_excelButton.clicked.connect(self.add_montage_data_in_bd_push)
        self.reportButton.clicked.connect(self.report_bd_push)
        self.dateEdit.setDate(QtCore.QDate.currentDate())
        self.dateEdit.setDisplayFormat('dd.MM.yyyy')
        self.filename = ''
        self.filetype = ''
        self.path_to_exel = ""
        self.data_montage_progress = []
        self.reportButton.clicked.connect(self.report_montage_bd_push)
        self.MainWindow = ""

    def back_btn_push(self):
        self.close()
        self.MainWindow = Main_menu()
        self.MainWindow.show()

    def choose_report_btn_push(self):
        self.data_montage_progress.clear()
        self.filename, self.filetype = QFileDialog.getOpenFileName(self,
                                                                   "Выбрать файл",
                                                                   ".",
                                                                   "Excel Files(*.xlsx);;All Files(*)")
        self.plainTextEdit.appendHtml("Путь: <b>{}</b> <br> <b>{:*^54}</b>"
                                      "".format(self.filename, self.filetype))
        self.path_to_exel = "{}".format(self.filename)
        if self.path_to_exel is None or self.path_to_exel == '':
            error = f'Файл не выбран!'
            self.MainWindow = ErrorAddReport(error)
            self.MainWindow.show()
        else:
            try:
                wb = load_workbook(f"{self.path_to_exel}", data_only=True)
                # Из базы данных берутся все активные проекты и используется в качестве имени листа
                for progect_data in db.get_all_active_montage_progect():
                    print(progect_data)
                    sheet_ranges_1 = wb[progect_data[0]]
            except:
                error = f'Файл не выбран! В данном файле не все активные проекты.'
                self.MainWindow = ErrorAddReport(error)
                self.MainWindow.show()

    def add_montage_data_in_bd_push(self):
        date = self.dateEdit.date().toString('dd MMM yy')
        real_date = self.dateEdit.date().toString('yyyy.MM.dd')
        try:
            wb = load_workbook(f"{self.path_to_exel}", data_only=True)
            # Из базы данных берутся все активные проекты и используется в качестве имени листа
            for progect_data in db.get_all_active_montage_progect():
                sheet_ranges_1 = wb[progect_data[0]]
                try:
                    res = str(db.get_id_progect(progect_data[0]))
                    key = str(int(re.search('\d+', res).group(0)) + 1)
                    db.delite_data_montag_report(key, real_date)
                    real_name_progect = progect_data[0]
                    organisation_work = round(sheet_ranges_1['C3'].value * 100, 2)
                    installation_of_metal_frame = round(sheet_ranges_1['D3'].value * 100, 2)
                    installation_of_fencing_constractions = round(sheet_ranges_1['F3'].value * 100, 2)
                    installation_of_engineering_system = round(sheet_ranges_1['H3'].value * 100, 2)
                    finishing_work = round(sheet_ranges_1['J3'].value * 100, 2)
                    common_perc_of_work = round(sheet_ranges_1['K3'].value * 100, 2)
                    problems = sheet_ranges_1['L6'].value
                    if problems is None:
                        problems = ''
                    way_to_solve_problems = sheet_ranges_1['M6'].value
                    if way_to_solve_problems is None:
                        way_to_solve_problems = ''
                    # Добавление в базу данных для каждого проекта значений
                    db.add_montag_everyday_report(real_name_progect, date, organisation_work,
                                                  installation_of_metal_frame,
                                                  installation_of_fencing_constractions,
                                                  installation_of_engineering_system,
                                                  finishing_work, common_perc_of_work, problems,
                                                  way_to_solve_problems, real_date, key)
                    self.success_2.setPixmap(QPixmap("images/dop/success.png"))
                    message_title = "Успешное портирование данных"
                    message_text = f"Плановые значения успешно изменены."
                    msg = MessageDialogWindow(message_title, message_text)
                    if msg.success_msg() == 1:
                        pass
                except:
                    error = f'Проект {progect_data[0]}. Нет данных в ячейках.' \
                            f' Проверьте ячейки: C3, D3, F3, H3, J3, K3, L6, M6!'
                    self.MainWindow = ErrorAddReport(error)
                    self.MainWindow.show()
        except:
            error = 'Что-то пошло не так. Попробуйте выбрать файл снова!'
            self.MainWindow = ErrorAddReport(error)
            self.MainWindow.show()

    def report_montage_bd_push(self):
        try:
            date = self.dateEdit.date().toString('dd MMM yy')
            date_for_plan = self.dateEdit.date().toString('MMMM yyyy')
            monatge_excel(date, date_for_plan)
        except:
            error = 'Ошибка создания отчета. Возможно файл уже открыт. Попробуйте снова.'
            self.MainWindow = ErrorAddReport(error)
            self.MainWindow.show()


class ErrorAddReport(QDialog):
    # Окно ошибок с изменяемым тектстом
    def __init__(self, data, parent=None, flag=Qt.Dialog):
        super().__init__(parent, flag)
        uic.loadUi('ui/errors/error_dialog_report.ui', self)
        self.setWindowModality(QtCore.Qt.ApplicationModal)
        self.text_error = data
        self.label_dscr_of_error.clear()
        self.label_dscr_of_error.setText(self.text_error)
        self.setFocusPolicy(QtCore.Qt.StrongFocus)

    def focusOutEvent(self, event):
        self.activateWindow()
        self.raise_()
        self.show()

    def ok_btn_press(self):
        self.close()


class PrintReport(QMainWindow):
    # Вывод отчета по данным в БД
    def __init__(self, parent=None, flag=Qt.Window):
        super().__init__(parent, flag)
        uic.loadUi('ui/report_print.ui', self)
        # Выводим отчет согласно настройкам меню
        self.option_menu = int(re.search('\d+', str(db.get_menu_main())).group(0))
        if self.option_menu == 2:
            # Отчет по производству
            self.print_reportButton.clicked.connect(self.report_bd_push)
        elif self.option_menu == 3:
            # Отчет по монтажам
            self.print_reportButton.clicked.connect(self.report_montage_bd_push)
            self.label_2.clear()
            self.label_2.setText('Выберите дату, за которую хотите получить производственный отчет')
            self.type_progBox.hide()
            self.label_3.clear()
        self.backButton.clicked.connect(self.back_btn_push)
        self.dateEdit.setDate(QtCore.QDate.currentDate())
        self.dateEdit.setDisplayFormat('dd.MM.yyyy')

    def back_btn_push(self):
        # Кнопка назад
        self.close()
        self.MainWindow = Main_menu()
        self.MainWindow.show()

    def report_bd_push(self):
        # Кнопка вывода отчета по производству
        try:
            type_progect = self.type_progBox.currentText()
            date = self.dateEdit.date().toString('dd MMM yy')
            date_for_plan = self.dateEdit.date().toString('MMMM yyyy')
            excel(type_progect, date, date_for_plan)
        except:
            error = 'Ошибка создания отчета. Возможно файл уже открыт. Попробуйте снова.'
            self.MainWindow = ErrorAddReport(error)
            self.MainWindow.show()

    def report_montage_bd_push(self):
        # Кнопка вывода отчета по монтажам
        try:
            date = self.dateEdit.date().toString('dd MMM yy')
            date_for_plan = self.dateEdit.date().toString('MMMM yyyy')
            monatge_excel(date, date_for_plan)
        except:
            error = 'Ошибка создания отчета. Возможно файл уже открыт. Попробуйте снова.'
            self.MainWindow = ErrorAddReport(error)
            self.MainWindow.show()


class MessageDialogWindow(QtWidgets.QMessageBox):
    # Сообщение - подтверждение
    def __init__(self, title, text):
        super().__init__()
        self.title = title
        self.text_message = text
        self.msg = QtWidgets.QMessageBox(self)
        self.msg.setFocus()
        self.msg.setStyleSheet("font: 75 12pt bold \"Times New Romadn\";")

    def conirm_message(self):
        self.msg.setWindowIcon(QIcon("images/dop/attantion.png"))
        self.msg.setWindowTitle(self.title)
        self.msg.setIcon(QtWidgets.QMessageBox.Question)
        self.msg.setIconPixmap(QPixmap("images/dop/attantion.png"))
        self.msg.setText(self.text_message)
        buttonAceptar = self.msg.addButton("Да", QtWidgets.QMessageBox.YesRole)
        buttonCancelar = self.msg.addButton("Отменить", QtWidgets.QMessageBox.RejectRole)
        self.msg.setDefaultButton(buttonAceptar)
        self.msg.exec_()
        if self.msg.clickedButton() == buttonAceptar:
            return 1
        else:
            return 0

    def success_msg(self):
        self.msg.setWindowIcon(QIcon("images/dop/success.png"))
        self.msg.setWindowTitle(self.title)
        self.msg.setIcon(QtWidgets.QMessageBox.Question)
        self.msg.setIconPixmap(QPixmap("images/dop/success.png"))
        self.msg.setText(self.text_message)
        buttonAceptar = self.msg.addButton("Ок", QtWidgets.QMessageBox.YesRole)
        self.msg.setDefaultButton(buttonAceptar)
        self.msg.exec_()
        if self.msg.clickedButton() == buttonAceptar:
            return 1


def application():
    app = QtWidgets.QApplication(sys.argv)
    app.setWindowIcon(QIcon('images/report.png'))
    option_menu = int(re.search('\d+', str(db.get_menu_main())).group(0))
    print(option_menu)
    if option_menu == 1:
        mainwindow = ChooseTypeOfMenu()
        mainwindow.show()
    else:
        mainwindow = Main_menu()
        mainwindow.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    application()
