import os
import re
import sys
from PyQt5.QtWidgets import QDialog
from PyQt5.QtCore import *
from datetime import datetime, date
from PyQt5 import QtCore, QtWidgets, QtGui
import xlsxwriter
from PyQt5 import uic
from openpyxl import load_workbook

from config import db


class ExcelWork:
    def __init__(self, t_progect, n_date, date_for_plan):
        super().__init__()
        self.type_prog = t_progect
        self.date_of_report = n_date
        self.data_of_mounth_plan = date_for_plan
        name_of_first_ws = ''
        lbl_for_plan = ''
        if self.type_prog == 'КМД':
            self.workbook = xlsxwriter.Workbook("Отчеты/Отчет по цеху Металлоконструкций.xlsx")
            lbl_for_plan = "План месяц, т"
            name_of_first_ws = 'Сводный отчет по цеху Металлоконструкций'
            self.workbook.set_properties({
                'title': f'Производственный отчет по производству КМД',
                'subject': 'With document properties',
                'author': 'Ivan Metliaev',
                'manager': '',
                'company': 'Тентовые конструкции',
                'category': 'КМД',
                'keywords': 'КМД, Ангары, Металл',
                'created': datetime.today(),
                'comments': 'Created with Python and Ivan Metliaev program'})

        elif self.type_prog == 'СПУ':
            self.workbook = xlsxwriter.Workbook(f'Отчеты/Отчет по участку изготовления СПУ.xlsx')
            name_of_first_ws = 'Сводный отчет по участку производства утеплителя'
            lbl_for_plan = "План месяц, м2"
            self.workbook.set_properties({
                'title': f'Производственный отчет по производству СПУ',
                'subject': 'With document properties',
                'author': 'Ivan Metliaev',
                'manager': '',
                'company': 'Тентовые конструкции',
                'category': 'СПУ',
                'keywords': 'СПУ, Утеплитель, Покрытие',
                'created': datetime.today(),
                'comments': 'Created with Python and Ivan Metliaev program'})
        elif self.type_prog == 'ПВХ':
            self.workbook = xlsxwriter.Workbook(f'Отчеты/Отчет изготовления тентового полотна.xlsx')
            name_of_first_ws = 'Сводный отчет по участку производства тентового полотна'
            lbl_for_plan = "План месяц, м2"
            self.workbook.set_properties({
                'title': f'Производственный отчет по производству ПВХ',
                'subject': 'With document properties',
                'author': 'Ivan Metliaev',
                'manager': '',
                'company': 'Тентовые конструкции',
                'category': 'ПВХ',
                'keywords': 'ПВХ, Тент, Покрытие',
                'created': datetime.today(),
                'comments': 'Created with Python and Ivan Metliaev program'})
        # Форматы format()
        self.percent_format = self.workbook.add_format(
            {'border': 1, 'num_format': '0.00%', 'align': 'left', 'valign': 'vcenter'})
        self.percent_format_for_plan = self.workbook.add_format(
            {'num_format': '0.00%', 'align': 'left', 'valign': 'vcenter'})
        self.name_format = self.workbook.add_format(
            {'border': 1, 'bold': True, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True})
        self.name_format_main = self.workbook.add_format(
            {'align': 'center', 'valign': 'vcenter', 'text_wrap': True})
        self.date_format = self.workbook.add_format(
            {'border': 1, 'text_wrap': True, 'num_format': 'dd MMM yy', 'align': 'center', 'valign': 'vcenter'})
        self.date_format_main = self.workbook.add_format(
            {'text_wrap': True, 'num_format': 'dd MMM yy', 'align': 'center', 'valign': 'vcenter'})
        self.special_numb = self.workbook.add_format(
            {'border': 1, 'num_format': '#0', 'align': 'center', 'valign': 'vcenter'})
        self.float_numb_w_board = self.workbook.add_format(
            {'border': 1, 'num_format': '#0.00', 'align': 'center', 'valign': 'vcenter'})
        self.numb_w_border = self.workbook.add_format(
            {'num_format': '#0.00', 'align': 'center', 'valign': 'vcenter'})
        # Форматы для объединнеых ячеек
        self.name_merge_format = self.workbook.add_format({
            'align': 'center',
            'valign': 'vcenter',
            'num_format': '#0',
        })
        self.name_merge_format_main = self.workbook.add_format({
            'align': 'center',
            'valign': 'vcenter',
            'num_format': '#0',
            'bold': True
        })
        self.name_merge_format_spec = self.workbook.add_format({
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'num_format': '#0',
            'bold': True
        })
        self.name_merge_format_spec_2 = self.workbook.add_format({
            'border': 1,
            'align': 'right',
            'valign': 'vcenter',
            'num_format': '#0',
            'bold': True
        })
        self.name_merge_format_2 = self.workbook.add_format({
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'num_format': '#0',
            'fg_color': '#DDEBF7'
        })
        # Создаваемые листы
        self.worksheet_0 = self.workbook.add_worksheet(f'Сводный отчет')
        # Размер колонок
        self.worksheet_0.set_column(0, 0, 14)
        self.worksheet_0.set_column(1, 1, 20)
        self.worksheet_0.set_column(2, 2, 14)
        self.worksheet_0.set_column(3, 3, 17)
        self.worksheet_0.set_column(4, 4, 19)
        self.worksheet_0.set_column(5, 5, 13)
        self.worksheet_0.set_column(6, 6, 14)
        # Записи Ход изготовления ангара
        self.worksheet_0.merge_range(0, 0, 0, 4, name_of_first_ws,
                                self.name_merge_format)
        self.worksheet_0.write("F1", "Текущая дата:", self.name_format_main)
        self.worksheet_0.write("G1", f"{self.date_of_report}", self.date_format_main)
        self.worksheet_0.write("A4", lbl_for_plan, self.name_format_main)
        if self.type_prog == 'КМД':
            self.report_excel_kmd()
        elif self.type_prog == 'СПУ':
            self.report_excel_spu()

        elif self.type_prog == 'ПВХ':
            self.report_excel_pvh()

    def report_excel_kmd(self):
        kmd_plan = 0.0
        try:
            for kmd in db.get_kmd_plan_for_mounth(self.data_of_mounth_plan):
                kmd_plan = kmd
        except:
            pass
        self.worksheet_0.write("B4", kmd_plan, self.numb_w_border)
        self.worksheet_0.write("A5", f"Выполнение плана", self.name_format_main)
        self.worksheet_0.write_formula("B5", f"Выполнение плана", self.name_format_main)
        row_name = ['№', 'Проект', 'По проекту, т', 'Изготовлено на текущ. момент, т', 'Производственная готовность']
        # Заголовки первой таблицы
        curnt_numb_row = 7
        num = 0
        self.worksheet_0.write_row(6, 0, row_name, self.name_format)
        curnt_special_row = 7
        for info in db.get_massa_progect():
            key = info[1]
            for progect_data in db.get_info_about_today_kmd_report(key, self.date_of_report):
                # №
                num += 1
                self.worksheet_0.write(curnt_special_row, 0, num, self.special_numb)
                # Проект
                self.worksheet_0.write(curnt_special_row, 1, progect_data[0], self.name_format)
                # По проекту
                self.worksheet_0.write(curnt_special_row, 2, progect_data[1] / 1000, self.float_numb_w_board)
                # Готовность
                self.worksheet_0.write(curnt_special_row, 3, progect_data[2] / 1000, self.float_numb_w_board)
                # Процент
                self.worksheet_0.write(curnt_special_row, 4, progect_data[3] / 100, self.percent_format)
                curnt_special_row += 1
            curnt_numb_row += 1
        self.worksheet_0.write_formula(curnt_special_row, 3, f'=SUM(D8:D{curnt_special_row})', self.float_numb_w_board)
        self.worksheet_0.write_formula("B5", f'=D{curnt_special_row + 1}/B4', self.percent_format_for_plan)
        self.worksheet_0.merge_range(curnt_special_row, 0, curnt_special_row, 2, f'Итого:',
                                self.name_merge_format_spec)
        self.worksheet_0.conditional_format(7, 4, curnt_special_row, 4, {'type': 'data_bar'})
        self.worksheet_0.ignore_errors()
        for progect_in_work in db.get_info_progect_kmd():
            worksheet_1 = self.workbook.add_worksheet(f''.join(progect_in_work))
            # Размер колонок
            for width_row in range(1, 6):
                if width_row == 1:
                    worksheet_1.set_column(3, width_row, 10)
                elif width_row == 5:
                    worksheet_1.set_column(5, width_row, 18)
                else:
                    worksheet_1.set_column(3, width_row, 14)
            # Записи Ход изготовления ангара
            worksheet_1.merge_range(0, 0, 0, 5, f'Производственный отчет нарастающим итогом на отчетную дату',
                                    self.name_merge_format_spec)
            worksheet_1.write(1, 0, "Проект:", self.name_format)
            worksheet_1.merge_range(1, 1, 1, 5, f' ,'.join(progect_in_work), self.name_merge_format)
            worksheet_1.merge_range(2, 0, 2, 5, f'Конструкции металлические и деталировка', self.name_merge_format_2)
            row_name = ['№', 'Дата', 'Заготовка', 'Сварка', 'Покраска', 'Общая готовность ангара']
            # Заголовки первой таблицы
            worksheet_1.write_row(3, 0, row_name, self.name_format)
            curnt_numb_row = 4
            num = 0
            for info in db.get_info_ab_day_report(''.join(progect_in_work)):
                num += 1
                # №
                worksheet_1.write(curnt_numb_row, 0, num, self.special_numb)
                # Дата
                worksheet_1.write(curnt_numb_row, 1, info[2], self.date_format)
                # % Заготовки
                worksheet_1.write(curnt_numb_row, 2, info[3] / 100, self.percent_format)
                # % Сварки
                worksheet_1.write(curnt_numb_row, 3, info[4] / 100, self.percent_format)
                # % Покраски
                worksheet_1.write(curnt_numb_row, 4, info[5] / 100, self.percent_format)
                # % Общий проц готовности
                worksheet_1.write(curnt_numb_row, 5, info[6] / 100, self.percent_format)
                curnt_numb_row += 1

            worksheet_1.conditional_format(4, 2, curnt_numb_row, 5, {'type': 'data_bar'})
        self.workbook.close()
        os.startfile(f'Отчеты\Отчет по цеху Металлоконструкций.xlsx')

        # ОТЧЕТ ДЛЯ СПУ

    def report_excel_spu(self):
        tent_plan = 0.0
        try:
            for tent in db.get_spu_plan_for_mounth(self.data_of_mounth_plan):
                tent_plan = tent
        except:
            pass
        self.worksheet_0.write("B4", tent_plan, self.numb_w_border)
        self.worksheet_0.write("A5", f"Выполнение плана", self.name_format_main)
        self.worksheet_0.write_formula("B5", f"Выполнение плана", self.name_format_main)
        row_name = ['№', 'Проект', 'По проекту, м2', 'Изготовлено на текущ. момент, м2',
                    'Производственная готовность', 'Готовность к отгрузке']
        # Заголовки первой таблицы
        curnt_numb_row = 7
        num = 0
        self.worksheet_0.write_row(6, 0, row_name, self.name_format)
        curnt_special_row = 7
        for info in db.get_squer_spu_progect():
            key = info[1]
            for progect_data in db.get_info_about_today_spu_report(key, self.date_of_report):
                # №
                num += 1
                self.worksheet_0.write(curnt_numb_row, 0, num, self.special_numb)
                # Проект
                self.worksheet_0.write(curnt_numb_row, 1, progect_data[0], self.name_format)
                # По проекту
                self.worksheet_0.write(curnt_numb_row, 2, progect_data[1] / 1000, self.float_numb_w_board)
                # Готовность
                self.worksheet_0.write(curnt_numb_row, 3, progect_data[2] / 1000, self.float_numb_w_board)
                # Процент готовности
                self.worksheet_0.write(curnt_numb_row, 4, progect_data[3] / 100, self.percent_format)
                # Процент отгрузки
                self.worksheet_0.write(curnt_numb_row, 4, progect_data[4] / 100, self.percent_format)
                curnt_special_row += 1

            curnt_numb_row += 1
        self.worksheet_0.write_formula(curnt_special_row, 3, f'=SUM(D8:D{curnt_special_row})', self.float_numb_w_board)
        self.worksheet_0.write_formula("B5", f'=D{curnt_special_row + 1}/B4', self.percent_format_for_plan)
        self.worksheet_0.merge_range(curnt_special_row, 0, curnt_special_row, 2, f'Итого:',
                                     self.name_merge_format_spec)
        self.worksheet_0.conditional_format(7, 4, curnt_special_row, 5, {'type': 'data_bar'})
        for progect_in_work in db.get_info_progect_spu():
            # Создаваемые листы
            worksheet_1 = self.workbook.add_worksheet(f''.join(progect_in_work))
            # Размер колонок
            for width_row in range(1, 9):
                if width_row == 1:
                    worksheet_1.set_column(3, width_row, 10)
                else:
                    worksheet_1.set_column(3, width_row, 15)
            # Записи Ход изготовления ангара
            worksheet_1.merge_range(0, 0, 0, 8, f'Производственный отчет нарастающим итогом на отчетную дату',
                                    self.name_merge_format_spec)
            worksheet_1.write(1, 0, "Проект:", self.name_format)
            worksheet_1.merge_range(1, 1, 1, 8, f' ,'.join(progect_in_work), self.name_merge_format)
            worksheet_1.merge_range(2, 0, 2, 8, f'Утеплитель', self.name_merge_format_2)
            row_name = ['№', 'Дата', 'Раскрой полипропилена',
                        'Раскрой ПУ',
                        'Наклейка синтепона',
                        'Сборка без клея',
                        'Пробивка люверс',
                        'Упаковано',
                        'Общая готовность утеплителя']
            # Заголовки первой таблицы
            worksheet_1.write_row(3, 0, row_name, self.name_format)
            curnt_numb_row = 4
            num = 0
            for info in db.get_info_ab_day_report_spu(f''.join(progect_in_work)):
                num += 1
                # №
                worksheet_1.write(curnt_numb_row, 0, num, self.special_numb)
                # Дата
                worksheet_1.write(curnt_numb_row, 1, info[2], self.date_format)
                # % Раскрой полипропилена
                worksheet_1.write(curnt_numb_row, 2, info[3] / 100, self.percent_format)
                # % Раскрой ПУ
                if info[4] is not None:
                    worksheet_1.write(curnt_numb_row, 3, info[4] / 100, self.percent_format)
                else:
                    worksheet_1.write(curnt_numb_row, 3, '-', self.percent_format)
                # % Наклейка
                worksheet_1.write(curnt_numb_row, 4, info[5] / 100, self.percent_format)
                # % Сборка
                if info[6] is not None:
                    worksheet_1.write(curnt_numb_row, 5, info[6] / 100, self.percent_format)
                else:
                    worksheet_1.write(curnt_numb_row, 5, '-', self.percent_format)
                # % Пробивка люверс
                worksheet_1.write(curnt_numb_row, 6, info[7] / 100, self.percent_format)
                # % Упаковка
                worksheet_1.write(curnt_numb_row, 7, info[8] / 100, self.percent_format)
                # % готовности утеплителя
                worksheet_1.write(curnt_numb_row, 8, info[9] / 100, self.percent_format)
                curnt_numb_row += 1
            worksheet_1.conditional_format(4, 2, curnt_numb_row, 8, {'type': 'data_bar'})
        self.workbook.close()
        os.startfile(f'Отчеты\Отчет по участку изготовления СПУ.xlsx')

        # ОТЧЕТ ДЛЯ ПВХ

    def report_excel_pvh(self):
        tent_plan = 0.0
        try:
            for tent in db.get_tent_plan_for_mounth(self.data_of_mounth_plan):
                tent_plan = tent
        except:
            pass
        self.worksheet_0.write("B4", tent_plan, self.numb_w_border)
        self.worksheet_0.write("A5", f"Выполнение плана", self.name_format_main)
        self.worksheet_0.write_formula("B5", f"Выполнение плана", self.name_format_main)
        row_name = ['№', 'Проект', 'По проекту, м2', 'Изготовлено на текущ. момент, м2', 'Производственная готовность',
                    'Готовность к отгрузке']
        # Заголовки первой таблицы
        curnt_numb_row = 7
        num = 0
        self.worksheet_0.write_row(6, 0, row_name, self.name_format)
        curnt_special_row = 7
        for info in db.get_squer_pvh_progect():
            key = info[1]
            for progect_data in db.get_info_about_today_pvh_report(key, self.date_of_report):
                # №
                num += 1
                self.worksheet_0.write(curnt_numb_row, 0, num, self.special_numb)
                # Проект
                self.worksheet_0.write(curnt_numb_row, 1, progect_data[0], self.name_format)
                # По проекту
                self.worksheet_0.write(curnt_numb_row, 2, progect_data[1] / 1000, self.float_numb_w_board)
                # Готовность
                self.worksheet_0.write(curnt_numb_row, 3, progect_data[2] / 1000, self.float_numb_w_board)
                # Процент готовности
                self.worksheet_0.write(curnt_numb_row, 4, progect_data[3] / 100, self.percent_format)
                # Процент отгрузки
                self.worksheet_0.write(curnt_numb_row, 4, progect_data[4] / 100, self.percent_format)
                curnt_special_row += 1

            curnt_numb_row += 1
        self.worksheet_0.write_formula(curnt_special_row, 3, f'=SUM(D8:D{curnt_special_row})', self.float_numb_w_board)
        self.worksheet_0.write_formula("B5", f'=D{curnt_special_row + 1}/B4', self.percent_format_for_plan)
        self.worksheet_0.merge_range(curnt_special_row, 0, curnt_special_row, 2, f'Итого:',
                                self.name_merge_format_spec)
        self.worksheet_0.conditional_format(7, 4, curnt_special_row, 5, {'type': 'data_bar'})

        for progect_in_work in db.get_info_progect_pvh():
            # Создаваемые листы
            worksheet_1 = self.workbook.add_worksheet(f''.join(progect_in_work))
            # Размер колонок
            for width_row in range(1, 11):
                if width_row == 1:
                    worksheet_1.set_column(3, width_row, 10)
                elif width_row == 10:
                    worksheet_1.set_column(3, width_row, 18)
                else:
                    worksheet_1.set_column(3, width_row, 15)
            # Записи Ход изготовления ангара
            worksheet_1.merge_range(0, 0, 0, 10, f'Производственный отчет за {date.today()}',
                                    self.name_merge_format_spec)
            worksheet_1.write(1, 0, "Проект:", self.name_format)
            worksheet_1.merge_range(1, 1, 1, 10, f' ,'.join(progect_in_work), self.name_merge_format)
            worksheet_1.merge_range(2, 0, 2, 10, f'ПВХ покрытие внешнее и внутренее', self.name_merge_format_2)
            row_name = ['№', 'Дата', 'Раскрой полотна', 'Раскрой карманов', 'Раскрой нащельников', 'Сварка карманов',
                        'Приварить карманы', 'Пришить второй слой', 'Приварить нащельники', 'Упаковка',
                        'Общая готовности ТП']
            # Заголовки первой таблицы
            worksheet_1.write_row(3, 0, row_name, self.name_format)
            curnt_numb_row = 4
            num = 0
            for info in db.get_info_ab_day_report_pvh(f' ,'.join(progect_in_work)):
                num += 1
                # №
                worksheet_1.write(curnt_numb_row, 0, num, self.special_numb)
                # Дата
                worksheet_1.write(curnt_numb_row, 1, info[2], self.date_format)
                # % Раскрой полотна
                worksheet_1.write(curnt_numb_row, 2, info[3] / 100, self.percent_format)
                # % Раскрой карманов
                worksheet_1.write(curnt_numb_row, 3, info[4] / 100, self.percent_format)
                # % Раскрой Нащельников
                worksheet_1.write(curnt_numb_row, 4, info[5] / 100, self.percent_format)
                # % Сварка карманов
                worksheet_1.write(curnt_numb_row, 5, info[6] / 100, self.percent_format)
                # % Приварить карманы
                worksheet_1.write(curnt_numb_row, 6, info[7] / 100, self.percent_format)
                # % Полоса второго слоя
                worksheet_1.write(curnt_numb_row, 7, info[8] / 100, self.percent_format)
                # % нащельники приварить
                worksheet_1.write(curnt_numb_row, 8, info[9] / 100, self.percent_format)
                # % упаковать
                worksheet_1.write(curnt_numb_row, 9, info[10] / 100, self.percent_format)
                # % готовности утеплителя
                worksheet_1.write(curnt_numb_row, 10, info[11] / 100, self.percent_format)
                curnt_numb_row += 1
            worksheet_1.conditional_format(4, 2, curnt_numb_row, 10, {'type': 'data_bar'})

        self.workbook.close()
        os.startfile(f'Отчеты\Отчет изготовления тентового полотна.xlsx')


class ExcelMontage:
    def __init__(self, n_date, date_for_plan):
        super().__init__()
        self.date_of_report = n_date
        self.data_of_mounth_plan = date_for_plan
        self.report_excel_montage()

    def report_excel_montage(self):
        #  ОТЧЕТ ДЛЯ КМД
        workbook = xlsxwriter.Workbook(f'Отчеты/Отчет по монтажным работам.xlsx')
        # Форматы format()
        percent_format = workbook.add_format(
            {'border': 1, 'num_format': '0.00%', 'align': 'left', 'valign': 'vcenter'})
        percent_format_for_plan = workbook.add_format(
            {'num_format': '0.00%', 'align': 'left', 'valign': 'vcenter'})
        name_format = workbook.add_format(
            {'border': 1, 'bold': True, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True})
        name_format_no_bold = workbook.add_format(
            {'border': 1, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True})
        name_format_main = workbook.add_format(
            {'align': 'center', 'valign': 'vcenter', 'text_wrap': True})
        date_format = workbook.add_format(
            {'border': 1, 'text_wrap': True, 'num_format': 'dd MMM yy', 'align': 'center', 'valign': 'vcenter'})
        date_format_main = workbook.add_format(
            {'text_wrap': True, 'num_format': 'dd MMM yy', 'align': 'center', 'valign': 'vcenter'})
        special_numb = workbook.add_format(
            {'border': 1, 'num_format': '#0', 'align': 'center', 'valign': 'vcenter'})
        float_numb_w_board = workbook.add_format(
            {'border': 1, 'num_format': '#0.00', 'align': 'center', 'valign': 'vcenter'})
        numb_w_border = workbook.add_format(
            {'num_format': '#0.00', 'align': 'center', 'valign': 'vcenter'})
        # Форматы для объединнеых ячеек
        name_merge_format = workbook.add_format({
            'align': 'center',
            'valign': 'vcenter',
            'num_format': '#0',
        })
        name_merge_format_main = workbook.add_format({
            'align': 'center',
            'valign': 'vcenter',
            'num_format': '#0',
            'bold': True
        })
        name_merge_format_spec = workbook.add_format({
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'num_format': '#0',
            'bold': True
        })
        name_merge_format_spec_2 = workbook.add_format({
            'border': 1,
            'align': 'right',
            'valign': 'vcenter',
            'num_format': '#0',
            'bold': True
        })
        name_merge_format_2 = workbook.add_format({
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'num_format': '#0',
            'fg_color': '#DDEBF7'
        })
        workbook.set_properties({
            'title': f'Производственный отчет',
            'subject': 'With document properties',
            'author': 'Ivan Metliaev',
            'manager': '',
            'company': 'Тентовые конструкции',
            'category': 'КМД',
            'keywords': 'КМД, Ангары, Металл',
            'created': datetime.today(),
            'comments': 'Created with Python and Ivan Metliaev program'})
        # Создаваемые листы
        worksheet_0 = workbook.add_worksheet(f'Сводный отчет')
        # Размер колонок
        worksheet_0.set_column(0, 0, 14)
        worksheet_0.set_column(1, 1, 20)
        worksheet_0.set_column(2, 2, 14)
        worksheet_0.set_column(3, 3, 17)
        worksheet_0.set_column(4, 4, 19)
        worksheet_0.set_column(5, 5, 13)
        worksheet_0.set_column(6, 6, 14)
        # Записи Ход изготовления ангара
        worksheet_0.merge_range(0, 0, 0, 4, f'Cводный отчет по Монтажным работам',
                                name_merge_format_main)
        worksheet_0.write("F1", "Текущая дата:", name_format_main)
        worksheet_0.write("G1", f"{self.date_of_report}", date_format_main)
        row_name = ['№', 'Проект', 'Производственная готовность']
        # Заголовки первой таблицы
        curnt_numb_row = 7
        num = 0
        worksheet_0.write_row(6, 0, row_name, name_format)
        curnt_special_row = 7
        for info in db.get_all_active_montage_progect():
            key = info[1]
            for progect_data in db.get_info_about_today_montage_report(key, self.date_of_report):
                # №
                num += 1
                worksheet_0.write(curnt_special_row, 0, num, special_numb)
                # Проект
                worksheet_0.write(curnt_special_row, 1, progect_data[0], name_format)
                # По проекту
                # worksheet_0.write(curnt_special_row, 2, progect_data[1] / 1000, float_numb_w_board)
                # Готовность
                # worksheet_0.write(curnt_special_row, 3, progect_data[2] / 1000, float_numb_w_board)
                # Процент
                worksheet_0.write(curnt_special_row, 2, progect_data[1] / 100, percent_format)
                curnt_special_row += 1
            curnt_numb_row += 1
        # worksheet_0.write_formula(curnt_special_row, 3, f'=SUM(D8:D{curnt_special_row})', float_numb_w_board)
        # worksheet_0.write_formula("B5", f'=D{curnt_special_row + 1}/B4', percent_format_for_plan)
        # worksheet_0.merge_range(curnt_special_row, 0, curnt_special_row, 2, f'Итого:',name_merge_format_spec)
        worksheet_0.conditional_format(7, 4, curnt_special_row, 4, {'type': 'data_bar'})
        worksheet_0.ignore_errors()
        for progect_in_work in db.get_info_progect_montage():
            worksheet_1 = workbook.add_worksheet(f''.join(progect_in_work))
            # Размер колонок
            worksheet_1.set_column(3, 1, 10)
            worksheet_1.set_column(3, 2, 14)
            worksheet_1.set_column(3, 3, 18)
            worksheet_1.set_column(3, 4, 18)
            worksheet_1.set_column(5, 5, 18)
            worksheet_1.set_column(6, 6, 18)
            worksheet_1.set_column(7, 7, 22)
            worksheet_1.set_column(8, 8, 40)
            worksheet_1.set_column(9, 9, 45)
            # Записи Ход изготовления ангара
            worksheet_1.merge_range(0, 0, 0, 9, f'Производственный отчет нарастающим итогом на отчетную дату',
                                    name_merge_format_spec)
            worksheet_1.write(1, 0, "Проект:", name_format)
            worksheet_1.merge_range(1, 1, 1, 9, f' ,'.join(progect_in_work), name_merge_format)
            worksheet_1.merge_range(2, 0, 2, 9, f'Монтажные работы', name_merge_format_2)
            row_name = ['№', 'Дата', 'Организационные работы', 'Монтаж металлокаркаса',
                        'Монтаж ограждающих конструкций', 'Монтаж инженерных систем', 'Завершающие работы',
                        'Общая готовность ангара', 'Выявленые проблемы', 'Решение проблемы']
            # Заголовки первой таблицы
            worksheet_1.write_row(3, 0, row_name, name_format)
            curnt_numb_row = 4
            num = 0
            for info in db.get_info_evr_report_montage(f''.join(progect_in_work)):
                num += 1
                # №
                worksheet_1.write(curnt_numb_row, 0, num, special_numb)
                # Дата
                worksheet_1.write(curnt_numb_row, 1, info[2], date_format)
                # % Организационные работы
                worksheet_1.write(curnt_numb_row, 2, info[3] / 100, percent_format)
                # % Монтаж металлокаркаса
                worksheet_1.write(curnt_numb_row, 3, info[4] / 100, percent_format)
                # % Монтаж ограждающих конструкций
                worksheet_1.write(curnt_numb_row, 4, info[5] / 100, percent_format)
                # % Монтаж инженерных систем'
                worksheet_1.write(curnt_numb_row, 5, info[6] / 100, percent_format)
                # % Завершающие работы
                worksheet_1.write(curnt_numb_row, 6, info[7] / 100, percent_format)
                # % Общая готовность ангара
                worksheet_1.write(curnt_numb_row, 7, info[8] / 100, percent_format)
                # Проблемы
                if info[9] is not None:
                    worksheet_1.write(curnt_numb_row, 8, info[9], name_format_no_bold)
                else:
                    worksheet_1.write(curnt_numb_row, 8, '', name_format_no_bold)
                # Решение
                if info[10] is not None:
                    worksheet_1.write(curnt_numb_row, 9, info[10], name_format_no_bold)
                else:
                    worksheet_1.write(curnt_numb_row, 9, '', name_format_no_bold)

                curnt_numb_row += 1

            worksheet_1.conditional_format(4, 2, curnt_numb_row, 7, {'type': 'data_bar'})
        workbook.close()
        os.startfile(f'Отчеты\Отчет по монтажным работам.xlsx')


class ExcelMontageReportWrite():
    def __init__(self, filename, date, real_date):
        super().__init__()
        self.path_to_filename = "{}".format(filename)
        self.date = date
        self.real_date = real_date
